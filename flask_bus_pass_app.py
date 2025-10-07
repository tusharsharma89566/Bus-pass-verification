from flask import Flask, request, jsonify, render_template, session, redirect, url_for, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import sqlite3
import pandas as pd
import os
import json
import csv
import io
import secrets
import logging
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
import hashlib
import time

app = Flask(__name__)

# Security Configuration
app.config['SECRET_KEY'] = secrets.token_urlsafe(32)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('logs', exist_ok=True)

# Rate limiting storage
rate_limit_storage = {}

# Database initialization
def init_db():
    conn = sqlite3.connect('bus_pass.db')
    cursor = conn.cursor()
    
    # Students table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT UNIQUE NOT NULL,
            student_name TEXT NOT NULL,
            course TEXT NOT NULL,
            stoppage TEXT NOT NULL,
            bus_no TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK (role IN ('admin', 'staff')),
            is_active INTEGER DEFAULT 1,
            last_login TIMESTAMP,
            failed_login_attempts INTEGER DEFAULT 0,
            locked_until TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Audit log table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            ip_address TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Session table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            session_token TEXT UNIQUE NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            ip_address TEXT,
            user_agent TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Insert admin user if not exists
    cursor.execute('SELECT * FROM users WHERE username = ?', ('admin',))
    if not cursor.fetchone():
        admin_hash = generate_password_hash('Admin@123!', method='pbkdf2:sha256:100000')
        cursor.execute('''
            INSERT INTO users (username, password_hash, role) 
            VALUES (?, ?, ?)
        ''', ('admin', admin_hash, 'admin'))
        app.logger.info('Default admin user created')
    
    conn.commit()
    conn.close()

def get_db_connection():
    conn = sqlite3.connect('bus_pass.db')
    conn.row_factory = sqlite3.Row
    return conn

# Security decorators
def rate_limit(max_requests=5, window=300):  # 5 requests per 5 minutes
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
            current_time = time.time()
            
            if client_ip not in rate_limit_storage:
                rate_limit_storage[client_ip] = []
            
            # Clean old requests
            rate_limit_storage[client_ip] = [
                req_time for req_time in rate_limit_storage[client_ip] 
                if current_time - req_time < window
            ]
            
            if len(rate_limit_storage[client_ip]) >= max_requests:
                app.logger.warning(f'Rate limit exceeded for IP: {client_ip}')
                return jsonify({'error': 'Too many requests. Try again later.'}), 429
            
            rate_limit_storage[client_ip].append(current_time)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or 'session_token' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        
        # Validate session token
        conn = get_db_connection()
        session_record = conn.execute('''
            SELECT * FROM user_sessions 
            WHERE user_id = ? AND session_token = ? AND expires_at > ?
        ''', (session['user_id'], session['session_token'], datetime.now())).fetchone()
        conn.close()
        
        if not session_record:
            session.clear()
            return jsonify({'error': 'Session expired'}), 401
        
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'admin':
            log_audit_event(session.get('user_id'), 'UNAUTHORIZED_ACCESS_ATTEMPT', 
                          f'Attempted to access admin function: {request.endpoint}')
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

def validate_input(data, required_fields):
    """Validate and sanitize input data"""
    errors = []
    
    for field in required_fields:
        if field not in data or not data[field] or not data[field].strip():
            errors.append(f'{field} is required')
    
    if 'student_id' in data and data['student_id']:
        if not re.match(r'^[a-zA-Z0-9_-]{3,20}$', data['student_id']):
            errors.append('Student ID must be 3-20 characters, alphanumeric, underscore, or dash only')
    
    if 'student_name' in data and data['student_name']:
        if len(data['student_name'].strip()) < 2 or len(data['student_name'].strip()) > 100:
            errors.append('Student name must be 2-100 characters')
        if not re.match(r'^[a-zA-Z\s\.]+$', data['student_name'].strip()):
            errors.append('Student name can only contain letters, spaces, and dots')
    
    if 'bus_no' in data and data['bus_no']:
        if not re.match(r'^[A-Z0-9-]{1,10}$', data['bus_no'].strip()):
            errors.append('Bus number must be 1-10 characters, uppercase letters, numbers, and dashes only')
    
    return errors

def log_audit_event(user_id, action, details=None):
    """Log audit events"""
    try:
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO audit_log (user_id, action, details, ip_address)
            VALUES (?, ?, ?, ?)
        ''', (user_id, action, details, request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)))
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.error(f'Failed to log audit event: {e}')

def create_session(user_id):
    """Create secure session"""
    session_token = secrets.token_urlsafe(32)
    expires_at = datetime.now() + app.config['PERMANENT_SESSION_LIFETIME']
    
    conn = get_db_connection()
    conn.execute('''
        INSERT INTO user_sessions (user_id, session_token, expires_at, ip_address, user_agent)
        VALUES (?, ?, ?, ?, ?)
    ''', (user_id, session_token, expires_at, 
          request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr),
          request.headers.get('User-Agent', '')))
    conn.commit()
    conn.close()
    
    return session_token

def cleanup_expired_sessions():
    """Clean up expired sessions"""
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM user_sessions WHERE expires_at < ?', (datetime.now(),))
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.error(f'Failed to cleanup sessions: {e}')

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return render_template('dashboard.html')
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
@rate_limit(max_requests=5, window=900)  # 5 attempts per 15 minutes
def login():
    try:
        data = request.get_json()
        username = data.get('username', '').strip().lower()
        password = data.get('password', '')
        role = data.get('role', '').strip().lower()

        if not username or not password or not role:
            return jsonify({'error': 'Username, password, and role required'}), 400

        # Input validation
        if not re.match(r'^[a-zA-Z0-9_]{3,30}$', username):
            return jsonify({'error': 'Invalid username format'}), 400

        if role not in ['admin', 'staff']:
            return jsonify({'error': 'Invalid role selected'}), 400

        conn = get_db_connection()
        user = conn.execute('''
            SELECT * FROM users WHERE LOWER(username) = ? AND is_active = 1
        ''', (username,)).fetchone()

        if not user:
            log_audit_event(None, 'LOGIN_FAILED', f'Username not found: {username}')
            return jsonify({'error': 'Invalid credentials'}), 401

        # Check if account is locked
        if user['locked_until'] and datetime.fromisoformat(user['locked_until']) > datetime.now():
            log_audit_event(user['id'], 'LOGIN_FAILED', 'Account locked')
            return jsonify({'error': 'Account locked. Try again later.'}), 401

        # Verify password
        if not check_password_hash(user['password_hash'], password):
            # Increment failed attempts
            failed_attempts = user['failed_login_attempts'] + 1
            locked_until = None

            if failed_attempts >= 5:
                locked_until = datetime.now() + timedelta(minutes=30)

            conn.execute('''
                UPDATE users SET failed_login_attempts = ?, locked_until = ?
                WHERE id = ?
            ''', (failed_attempts, locked_until, user['id']))
            conn.commit()

            log_audit_event(user['id'], 'LOGIN_FAILED', f'Invalid password, attempts: {failed_attempts}')
            return jsonify({'error': 'Invalid credentials'}), 401

        # Verify role matches
        if user['role'] != role:
            log_audit_event(user['id'], 'LOGIN_FAILED', f'Role mismatch: selected {role}, actual {user["role"]}')
            return jsonify({'error': 'Invalid role for this user'}), 401

        # Successful login
        session_token = create_session(user['id'])

        # Reset failed attempts and update last login
        conn.execute('''
            UPDATE users SET failed_login_attempts = 0, locked_until = NULL, last_login = ?
            WHERE id = ?
        ''', (datetime.now(), user['id']))
        conn.commit()
        conn.close()

        # Set session
        session.permanent = True
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        session['session_token'] = session_token

        log_audit_event(user['id'], 'LOGIN_SUCCESS', f'Role: {user["role"]}')

        return jsonify({
            'success': True,
            'role': user['role'],
            'username': user['username']
        })

    except Exception as e:
        app.logger.error(f'Login error: {e}')
        return jsonify({'error': 'Login failed'}), 500

@app.route('/api/logout', methods=['POST'])
@login_required
def logout():
    try:
        # Remove session from database
        if 'session_token' in session:
            conn = get_db_connection()
            conn.execute('DELETE FROM user_sessions WHERE session_token = ?', 
                        (session['session_token'],))
            conn.commit()
            conn.close()
        
        log_audit_event(session.get('user_id'), 'LOGOUT', 'User logged out')
        session.clear()
        cleanup_expired_sessions()
        
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f'Logout error: {e}')
        return jsonify({'error': 'Logout failed'}), 500

@app.route('/api/students', methods=['GET'])
@login_required
def get_students():
    try:
        conn = get_db_connection()
        students = conn.execute('''
            SELECT student_id, student_name, course, stoppage, bus_no, created_at
            FROM students ORDER BY student_name
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(student) for student in students])
    except Exception as e:
        app.logger.error(f'Get students error: {e}')
        return jsonify({'error': 'Failed to fetch students'}), 500

@app.route('/api/students', methods=['POST'])
@login_required
@admin_required
def add_student():
    try:
        data = request.get_json()

        # Validate input
        required_fields = ['student_id', 'student_name', 'course', 'stoppage', 'bus_no']
        errors = validate_input(data, required_fields)

        if errors:
            return jsonify({'errors': errors}), 400

        # Sanitize data
        student_data = {
            'student_id': data['student_id'].strip(),
            'student_name': data['student_name'].strip().title(),
            'course': data['course'].strip().upper(),
            'stoppage': data['stoppage'].strip().title(),
            'bus_no': data['bus_no'].strip().upper()
        }

        conn = get_db_connection()

        # Check for duplicate student ID
        existing = conn.execute('SELECT id FROM students WHERE student_id = ?',
                              (student_data['student_id'],)).fetchone()
        if existing:
            return jsonify({'error': 'Student ID already exists'}), 400

        # Insert student
        conn.execute('''
            INSERT INTO students (student_id, student_name, course, stoppage, bus_no)
            VALUES (?, ?, ?, ?, ?)
        ''', (student_data['student_id'], student_data['student_name'],
              student_data['course'], student_data['stoppage'], student_data['bus_no']))

        conn.commit()
        conn.close()

        log_audit_event(session['user_id'], 'STUDENT_ADDED',
                       f'Student ID: {student_data["student_id"]}')

        return jsonify({'success': True, 'message': 'Student added successfully'})

    except Exception as e:
        app.logger.error(f'Add student error: {e}')
        return jsonify({'error': 'Failed to add student'}), 500

@app.route('/api/students/<student_id>', methods=['PUT'])
@login_required
@admin_required
def update_student(student_id):
    try:
        data = request.get_json()

        # Validate input
        required_fields = ['student_name', 'course', 'stoppage', 'bus_no']
        errors = validate_input(data, required_fields)

        if errors:
            return jsonify({'errors': errors}), 400

        # Sanitize data
        student_data = {
            'student_name': data['student_name'].strip().title(),
            'course': data['course'].strip().upper(),
            'stoppage': data['stoppage'].strip().title(),
            'bus_no': data['bus_no'].strip().upper()
        }

        conn = get_db_connection()

        # Check if student exists
        existing = conn.execute('SELECT id FROM students WHERE student_id = ?',
                              (student_id,)).fetchone()
        if not existing:
            return jsonify({'error': 'Student not found'}), 404

        # Update student
        conn.execute('''
            UPDATE students SET student_name = ?, course = ?, stoppage = ?, bus_no = ?, updated_at = CURRENT_TIMESTAMP
            WHERE student_id = ?
        ''', (student_data['student_name'], student_data['course'],
              student_data['stoppage'], student_data['bus_no'], student_id))

        conn.commit()
        conn.close()

        log_audit_event(session['user_id'], 'STUDENT_UPDATED',
                       f'Student ID: {student_id}')

        return jsonify({'success': True, 'message': 'Student updated successfully'})

    except Exception as e:
        app.logger.error(f'Update student error: {e}')
        return jsonify({'error': 'Failed to update student'}), 500

@app.route('/api/students/<student_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_student(student_id):
    try:
        conn = get_db_connection()

        # Check if student exists
        existing = conn.execute('SELECT id FROM students WHERE student_id = ?',
                              (student_id,)).fetchone()
        if not existing:
            return jsonify({'error': 'Student not found'}), 404

        # Delete student
        conn.execute('DELETE FROM students WHERE student_id = ?', (student_id,))

        conn.commit()
        conn.close()

        log_audit_event(session['user_id'], 'STUDENT_DELETED',
                       f'Student ID: {student_id}')

        return jsonify({'success': True, 'message': 'Student deleted successfully'})

    except Exception as e:
        app.logger.error(f'Delete student error: {e}')
        return jsonify({'error': 'Failed to delete student'}), 500

@app.route('/api/students/search', methods=['POST'])
@login_required
def search_student():
    try:
        data = request.get_json()
        student_id = data.get('student_id', '').strip()
        
        if not student_id:
            return jsonify({'error': 'Student ID required'}), 400
        
        conn = get_db_connection()
        student = conn.execute('''
            SELECT student_id, student_name, course, stoppage, bus_no
            FROM students WHERE LOWER(student_id) = LOWER(?)
        ''', (student_id,)).fetchone()
        conn.close()
        
        if student:
            log_audit_event(session['user_id'], 'STUDENT_SEARCHED', 
                           f'Student ID: {student_id}')
            return jsonify({'success': True, 'student': dict(student)})
        else:
            return jsonify({'success': False, 'message': 'Student not found'})
            
    except Exception as e:
        app.logger.error(f'Search student error: {e}')
        return jsonify({'error': 'Search failed'}), 500

@app.route('/api/students/upload', methods=['POST'])
@login_required
@admin_required
def upload_students():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        filename = secure_filename(file.filename)
        if not filename.lower().endswith(('.csv', '.xlsx')):
            return jsonify({'error': 'Only CSV and Excel files allowed'}), 400
        
        # Save uploaded file temporarily
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Read file based on extension
            if filename.lower().endswith('.csv'):
                df = pd.read_csv(filepath)
            else:
                df = pd.read_excel(filepath)
            
            # Validate required columns
            required_cols = ['student_id', 'student_name', 'course', 'stoppage', 'bus_no']
            if not all(col in df.columns for col in required_cols):
                return jsonify({'error': f'Missing required columns: {required_cols}'}), 400
            
            # Clean and validate data
            df = df.dropna(subset=required_cols)
            df = df[required_cols]  # Keep only required columns
            
            success_count = 0
            error_count = 0
            errors = []
            
            conn = get_db_connection()
            
            for index, row in df.iterrows():
                try:
                    # Validate each row
                    row_data = {
                        'student_id': str(row['student_id']).strip(),
                        'student_name': str(row['student_name']).strip().title(),
                        'course': str(row['course']).strip().upper(),
                        'stoppage': str(row['stoppage']).strip().title(),
                        'bus_no': str(row['bus_no']).strip().upper()
                    }
                    
                    validation_errors = validate_input(row_data, required_cols)
                    if validation_errors:
                        error_count += 1
                        errors.append(f'Row {index + 1}: {", ".join(validation_errors)}')
                        continue
                    
                    # Try to insert (skip duplicates)
                    conn.execute('''
                        INSERT OR IGNORE INTO students 
                        (student_id, student_name, course, stoppage, bus_no)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (row_data['student_id'], row_data['student_name'],
                          row_data['course'], row_data['stoppage'], row_data['bus_no']))
                    
                    if conn.total_changes > 0:
                        success_count += 1
                    else:
                        error_count += 1
                        errors.append(f'Row {index + 1}: Duplicate student ID')
                        
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {index + 1}: {str(e)}')
            
            conn.commit()
            conn.close()
            
            log_audit_event(session['user_id'], 'BULK_UPLOAD', 
                           f'Success: {success_count}, Errors: {error_count}')
            
            return jsonify({
                'success': True,
                'message': f'Upload completed. {success_count} students added, {error_count} errors.',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # Limit error messages
            })
            
        finally:
            # Clean up uploaded file
            if os.path.exists(filepath):
                os.remove(filepath)
                
    except Exception as e:
        app.logger.error(f'Upload error: {e}')
        return jsonify({'error': 'Upload failed'}), 500

@app.route('/api/students/export/<format>')
@login_required
@admin_required
def export_students(format):
    try:
        conn = get_db_connection()
        students = conn.execute('''
            SELECT student_id, student_name, course, stoppage, bus_no, created_at
            FROM students ORDER BY student_name
        ''').fetchall()
        conn.close()
        
        if not students:
            return jsonify({'error': 'No data to export'}), 400
        
        # Convert to list of dicts
        data = [dict(student) for student in students]
        
        if format.lower() == 'csv':
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
            
            response = app.response_class(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment; filename=students.csv'}
            )
            
        elif format.lower() == 'xlsx':
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, student in enumerate(data, 2):
                for col, value in enumerate(student.values(), 1):
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(output)
            output.seek(0)
            
            response = app.response_class(
                output.read(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename=students.xlsx'}
            )
        else:
            return jsonify({'error': 'Invalid format'}), 400
        
        log_audit_event(session['user_id'], 'DATA_EXPORT', f'Format: {format.upper()}')
        return response
        
    except Exception as e:
        app.logger.error(f'Export error: {e}')
        return jsonify({'error': 'Export failed'}), 500

@app.route('/api/audit-logs')
@login_required
@admin_required
def get_audit_logs():
    try:
        conn = get_db_connection()
        logs = conn.execute('''
            SELECT al.*, u.username 
            FROM audit_log al
            LEFT JOIN users u ON al.user_id = u.id
            ORDER BY al.timestamp DESC
            LIMIT 100
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(log) for log in logs])
    except Exception as e:
        app.logger.error(f'Audit logs error: {e}')
        return jsonify({'error': 'Failed to fetch audit logs'}), 500

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File too large'}), 413

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(e):
    app.logger.error(f'Internal server error: {e}')
    return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=False, host='0.0.0.0', port=5000)